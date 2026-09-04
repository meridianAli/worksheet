#!/usr/bin/env python3
"""Hold a spoken script to the changes its analyst would have to make.

The question is not "does the script mention this cell" - it is "reading only
this script, with the input workbook open in front of you, would you have made
this change?" That splits deterministically on where the vocabulary comes from:

  MUST_SAY_VALUE  a literal number in the output that appears nowhere in the
                  input workbook. The analyst cannot derive it or read it off
                  the source file, so the senior had to say it. Missing = FAIL.

  MUST_NAME       a label/heading in the output that appears nowhere in the
                  input workbook. New vocabulary - a tab name, a section title,
                  a metric the model did not previously carry. The senior had
                  to name it. Missing = FAIL.

  DERIVABLE       everything whose vocabulary is already in the input workbook
                  (existing line items, existing headers) plus every formula
                  behind a named deliverable. A competent analyst produces
                  these from the source file and the named ask, so the script
                  is not required to dictate them. Reported, never enforced.

That last bucket is the point. A senior does not read out 400 rows of an
amortization schedule or re-read line items the analyst can already see; the
test must not demand it. What it does demand is that nothing arrives out of
thin air.

Usage:
  python3 coverage.py --diff changes.json --input <input.xlsx> --script script.md
"""
import argparse, gzip, json, os, re, sys
from openpyxl import load_workbook

LEXICON = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'standard_lexicon.txt')

# Spoken forms a script may legitimately use for a written label.
ALIASES = {
    '&': ' and ', '%': ' percent ', '#': ' number ', '/': ' ', '(': ' ', ')': ' ',
    "'": '', '"': '', ',': '', ':': ' ', ';': ' ', '-': ' ', '_': ' ', '.': ' ',
    '$': ' ', '+': ' ', '?': ' ', '!': ' ', '\u2014': ' ', '\u2013': ' ',
    '*': ' ', '`': ' ', '|': ' ', '\u2212': '-',
}
# Words that carry no instruction - dropped before comparing a label to speech.
FILLER = set('''the a an and or of to in for on at as is are be that this it its
with from by all each both not no use using into then also same new should
which where when what s t re ll m'''.split())
# Written shorthand -> what a person says out loud.
SPOKEN = [
    (r'\bnoso\b', 'noso'), (r'\bwaso\b', 'waso'), (r'\bbop\b', 'bop'),
    (r'\beop\b', 'eop'), (r'\bi/o\b', 'io'), (r'\bd&a\b', 'd and a'),
    (r'\$mm\b', 'mm'), (r"\$'000s", '000s'),
]
NUMISH = re.compile(r'-?\d[\d,]*\.?\d*')


UNIT_FRAGMENT = re.compile(
    r'\b(mm|000s|m|k|bn|usd|percent|of bop|of opex|of revenue|of operating expenses'
    r'|excl cash|excl debt|non operating|pre money|post money|ipo)\b')


def norm(s):
    s = str(s).replace('\u0394', ' delta ').lower()
    s = re.sub(r'\bpro\s*forma\b', 'proforma', s)
    # fy24a and fy2024a are the same year said two ways
    s = re.sub(r'\bfy(\d{2})([ap]?)\b', lambda m: f'fy20{m.group(1)}{m.group(2)}', s)
    s = re.sub(r'\bdelta\b', 'change', s)
    for a, b in ALIASES.items():
        s = s.replace(a, b)
    for pat, rep in SPOKEN:
        s = re.sub(pat, rep, s)
    s = re.sub(r'^[\s+/-]+', '', s)
    s = re.sub(r'^\s*(change in|chg)\s+', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def stem(w):
    return w[:-1] if len(w) > 3 and w.endswith('s') else w


def tokens(s):
    return [stem(w) for w in norm(s).split() if w not in FILLER and len(w) > 1]


def spoken_covers(label, script_tokens, spoken):
    """A senior paraphrases. "Offset - that's when extra principal happens"
    covers the header "Offset (when extra principal happens)", and "shareholder
    dilution at different share issuance" covers "...at different issuance of
    shares". So a label counts as said when the script contains every
    instruction-carrying word in it, not when it contains the string."""
    n = norm(label)
    if n and n in spoken:
        return True
    t = tokens(label)
    if not t:
        return True          # units-only or punctuation-only cell
    if all(w in script_tokens for w in t):
        return True
    # Fall back to the head of the label: a parenthetical is a qualifier on
    # something already asked for ("Base (invisible)" is the waterfall's base
    # series), not a separate instruction.
    head = re.sub(r'\([^)]*\)', ' ', str(label))
    ht = tokens(head)
    return bool(ht) and all(w in script_tokens for w in ht)


def lexicon():
    out = set()
    for line in open(LEXICON):
        line = line.split('#')[0].strip()
        if line:
            out.add(norm(line))
    return out


def is_standard(label, lex):
    n = norm(label)
    if n in lex:
        return True
    bare = UNIT_FRAGMENT.sub(' ', n)
    bare = re.sub(r'\s+', ' ', bare).strip()
    return bool(bare) and bare in lex


def numkey(v):
    """Compare numbers by value, so 42.0 == 42 and 1,282 == 1282."""
    try:
        f = float(str(v).replace(',', '').rstrip('%'))
    except (TypeError, ValueError):
        return None
    return f'{f:.6f}'.rstrip('0').rstrip('.')


def workbook_vocab(path):
    """Every label and every number already visible in the input workbook."""
    wb = load_workbook(path, data_only=False, read_only=True)
    wbv = load_workbook(path, data_only=True, read_only=True)
    labels, numbers = set(), set()
    for name in wb.sheetnames:
        labels.add(norm(name))
        for src in (wb[name], wbv[name]):
            for row in src.iter_rows():
                for c in row:
                    v = c.value
                    if v is None:
                        continue
                    if isinstance(v, str):
                        if v.startswith('='):
                            for m in NUMISH.findall(v):
                                k = numkey(m)
                                if k:
                                    numbers.add(k)
                            continue
                        labels.add(norm(v))
                    else:
                        k = numkey(v)
                        if k:
                            numbers.add(k)
    wb.close(); wbv.close()
    return labels, numbers


RANGE = re.compile(
    r'(-?[\d,]*\.?\d+)\s*(?:to|through|-|\u2013)\s*(-?[\d,]*\.?\d+)[^.]{0,60}?'
    r'(?:(?:in|by)\s+)?(?:increments?\s+of|steps?\s+of|by)\s+(-?[\d,]*\.?\d+)', re.I)


# A senior says "zero", "nil", "a hundred and five" - not always digits.
WORD_NUM = {
    'zero': 0, 'nil': 0, 'none': 0, 'one': 1, 'two': 2, 'three': 3, 'four': 4,
    'five': 5, 'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10,
    'eleven': 11, 'twelve': 12, 'twenty': 20, 'thirty': 30, 'forty': 40,
    'fifty': 50, 'hundred': 100, 'thousand': 1000, 'million': 1000000,
}


def script_index(path):
    text = open(path).read()
    # markdown scaffolding is not speech: drop heading markers and list bullets
    text = re.sub(r'^\s{0,3}#{1,6}\s+', '', text, flags=re.M)
    text = re.sub(r'^\s*[-*]\s+', '', text, flags=re.M)
    spoken = norm(text)
    nums = set()
    for w, v in WORD_NUM.items():
        if re.search(rf'\b{w}\b', spoken):
            nums.add(numkey(v))
    for m in NUMISH.findall(text):
        k = numkey(m)
        if k:
            nums.add(k)
    # A senior states a grid as "160 to 200 in increments of 10" and never reads
    # out 170 and 190. Expand any stated range so its interior counts as said.
    grids = []
    for lo, hi, step in RANGE.findall(text.replace('$', '')):
        try:
            lo, hi, step = float(lo.replace(',', '')), float(hi.replace(',', '')), float(step.replace(',', ''))
        except ValueError:
            continue
        if step <= 0 or (hi - lo) / step > 500:
            continue
        v, n = lo, 0
        while v <= hi + step / 1000:
            nums.add(numkey(v))
            v += step
            n += 1
        grids.append((step, n))
    return spoken, nums, grids


def value_covered(k, script_nums):
    """A spoken number matches its cell sign-insensitively and in either
    percent form: a senior says "1.1%" for a cell holding 0.011, and
    "repayments of 0.5" for a cell holding -0.5."""
    try:
        f = float(k)
    except (TypeError, ValueError):
        return False
    for cand in (f, -f, f * 100, -f * 100, f / 100, -f / 100):
        if numkey(cand) in script_nums:
            return True
    return False


def rowlabel(changes, sheet, row):
    for c in changes:
        if c['sheet'] == sheet and c.get('row') == row and c['type'] == 'label' \
                and c.get('col') in (1, 2):
            return c['after']['value']
    return None


def ch_pre(diff):
    return diff['changes']


def deliberate(c, formula_rows):
    """A typed value the analyst could not have arrived at by working.

    Two shapes matter beyond "number not in the source file": a literal
    written as a formula ('=60000000' - somebody parked that on purpose), and
    a number typed into a row that is otherwise formula-driven (an override).
    Both are decisions, and a decision has to come from the person briefing."""
    if c['after'].get('as_formula'):
        return True
    return formula_rows.get((c['sheet'], c.get('row')), 0) >= 2


STEP_FORMULA = re.compile(
    r'^=\s*\$?[A-Z]{1,3}\$?\d+\s*([+-])\s*([\d,]*\.?\d+)\s*$')


def grid_steps(changes):
    """Steps the senior chose, read out of the formulas that walk a grid axis.

    A sensitivity axis is built as '=G80+2' off a typed anchor, so the increment
    never appears as a hardcoded cell - but "in $2.00 increments" is exactly the
    kind of thing a senior says and an analyst cannot invent. The literal offset
    in a neighbour-stepping formula is that increment.
    """
    steps, anchors = {}, {}
    for c in changes:
        if c.get('row') is None:
            continue
        if c['type'] == 'hardcode':
            # the typed value the axis steps away from, in both directions
            anchors.setdefault((c['sheet'], 'row', c['row']), set()).add(c['col'])
            anchors.setdefault((c['sheet'], 'col', c['col']), set()).add(c['row'])
            continue
        if c['type'] != 'formula':
            continue
        m = STEP_FORMULA.match(str(c['after'].get('formula', '')))
        if not m:
            continue
        step = abs(float(m.group(2).replace(',', '')))
        steps.setdefault((c['sheet'], 'row', c['row'], step), {})[c['col']] = c['cell']
        steps.setdefault((c['sheet'], 'col', c['col'], step), {})[c['row']] = c['cell']

    out = {}
    for (sheet, kind, idx, step), cells in steps.items():
        anc = anchors.get((sheet, kind, idx), set())
        pos = sorted(cells)
        run = [pos[0]]
        for p in pos[1:]:
            # a gap of one is the anchor cell sitting inside the axis
            if p - run[-1] <= 2:
                run.append(p)
            else:
                out.update(_axis(sheet, kind, run, anc, step, cells))
                run = [p]
        out.update(_axis(sheet, kind, run, anc, step, cells))
    return out


def _axis(sheet, kind, run, anchors, step, cells):
    lo, hi = run[0], run[-1]
    if lo - 1 in anchors:
        lo -= 1
    if hi + 1 in anchors:
        hi += 1
    return {(sheet, kind, cells[run[0]], step): (hi - lo + 1, cells[run[0]])}


def audit(diff, in_labels, in_numbers, spoken, script_nums, lex, script_grids):
    reqs = []
    stoks = set(tokens(spoken))
    # rows that carry formula work, to spot a value typed in among them
    formula_rows = {}
    for c in ch_pre(diff):
        if c['type'] == 'formula' and c.get('row'):
            formula_rows[(c['sheet'], c['row'])] = formula_rows.get((c['sheet'], c['row']), 0) + 1
    ch = diff['changes']
    for c in ch:
        if c['type'] == 'hardcode':
            k = numkey(c['after']['value'])
            if k is None or (k in in_numbers and not deliberate(c, formula_rows)):
                reqs.append(dict(tier='DERIVABLE', why='value already in input',
                                 sheet=c['sheet'], cell=c['cell'], item=c['after']['value']))
                continue
            reqs.append(dict(tier='MUST_SAY_VALUE', sheet=c['sheet'], cell=c['cell'],
                             item=c['after']['value'], key=k,
                             label=rowlabel(ch, c['sheet'], c.get('row')),
                             covered=value_covered(k, script_nums)))
        elif c['type'] == 'label':
            n = norm(c['after']['value'])
            if not n or n in in_labels:
                reqs.append(dict(tier='DERIVABLE', why='label already in input',
                                 sheet=c['sheet'], cell=c['cell'], item=c['after']['value']))
                continue
            v = c['after']['value']
            if c.get('col') == 1 and (c.get('row', 99) <= 2 or
                                      (v.isupper() and len(v.split()) > 1)):
                reqs.append(dict(tier='DERIVABLE', why='sheet title / section banner',
                                 sheet=c['sheet'], cell=c['cell'], item=v))
                continue
            if is_standard(c['after']['value'], lex):
                reqs.append(dict(tier='DERIVABLE', why='standard finance vocabulary',
                                 sheet=c['sheet'], cell=c['cell'], item=c['after']['value']))
                continue
            reqs.append(dict(tier='MUST_NAME', sheet=c['sheet'], cell=c['cell'],
                             item=c['after']['value'],
                             covered=spoken_covers(c['after']['value'], stoks, spoken)))
        else:
            reqs.append(dict(tier='DERIVABLE', why=c['type'], sheet=c['sheet'],
                             cell=c['cell'], item=c.get('after', {}).get('formula')))
    for (sheet, kind, idx, step), (points, cell) in grid_steps(ch).items():
        if step in (0.0, 1.0) or points < 3:
            continue          # a +1 walk is a row or period counter, not a choice
        # The stated range has to have the right step AND the right extent, so
        # "160 to 200 by 10" passes and "160 to 190 by 10" does not.
        ok = any(abs(st - step) < 1e-9 and n == points for st, n in script_grids)
        reqs.append(dict(tier='MUST_SAY_RANGE', sheet=sheet, cell=cell,
                         item=f'{points}-point axis, step {step:g}', covered=ok))
    for s in diff['sheets_added']:
        # A tab name is an identifier, not a description. "Revenue_Forecast" is
        # not covered by having said "forecast" somewhere - the analyst has to
        # be told what to call the tab, exactly.
        reqs.append(dict(tier='MUST_NAME', sheet=s, cell=None, item=f'tab: {s}',
                         covered=norm(s) in spoken or
                                 norm(s).replace(' ', '') in spoken.replace(' ', '')))
    return reqs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--diff', required=True)
    ap.add_argument('--input', required=True)
    ap.add_argument('--script', required=True)
    ap.add_argument('--json')
    ap.add_argument('--quiet', action='store_true')
    a = ap.parse_args()

    op = gzip.open if a.diff.endswith('.gz') else open
    diff = json.load(op(a.diff, 'rt'))
    in_labels, in_numbers = workbook_vocab(a.input)
    spoken, script_nums, script_grids = script_index(a.script)
    reqs = audit(diff, in_labels, in_numbers, spoken, script_nums, lexicon(),
                 script_grids)

    enforced = [r for r in reqs if r['tier'] != 'DERIVABLE']
    # One requirement per distinct thing, not per cell.
    seen, uniq = set(), []
    for r in enforced:
        k = (r['tier'], norm(r['item']))
        if k in seen:
            continue
        seen.add(k); uniq.append(r)
    missed = [r for r in uniq if not r['covered']]

    print(f"{a.script.split('/')[-1]}  vs  {diff['output'].split('/')[-1]}")
    print(f"  {len(diff['changes'])} cell changes -> {len(uniq)} enforced requirements "
          f"({len(reqs) - len(enforced)} derivable, not enforced)")
    for tier in ('MUST_SAY_VALUE', 'MUST_SAY_RANGE', 'MUST_NAME'):
        t = [r for r in uniq if r['tier'] == tier]
        m = [r for r in t if not r['covered']]
        print(f"  {tier:15s} {len(t) - len(m):3d}/{len(t):3d} covered")
        if m and not a.quiet:
            for r in m[:40]:
                loc = f"{r['sheet']}!{r['cell']}" if r['cell'] else r['sheet']
                extra = f"  (row: {r['label']})" if r.get('label') else ''
                print(f"      MISS {loc:28s} {str(r['item'])[:52]}{extra}")
            if len(m) > 40:
                print(f"      ... {len(m) - 40} more")
    if a.json:
        json.dump(reqs, open(a.json, 'w'), indent=1, default=str)
    return 1 if missed else 0


if __name__ == '__main__':
    sys.exit(main())
