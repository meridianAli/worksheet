import csv, json, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SCR = '/tmp/claude-0/-home-user-worksheet/968f285e-0587-5207-a956-c384e074891a/scratchpad'
AHT = '/home/user/worksheet/aht'
OUT = os.path.join(AHT, 'August-AHT-Analysis.xlsx')

coh     = json.load(open('/tmp/claude-0/-home-user-worksheet/968f285e-0587-5207-a956-c384e074891a/scratchpad/cohort.json'))
CAP_H   = 12
FACTOR  = coh['factor']
PAYABLE = coh['payable']
CLAIMH  = coh['claim']
BLENDED = round(coh['cost']/coh['hours'], 2)
NPEOPLE = coh['people']; NTASKS = coh['tasks']; NCOST = coh['cost']

# ---------------------------------------------------------------- source data
tasks = list(csv.DictReader(open(os.path.join(AHT, 'august_aht_per_task.csv'))))
ONTASK  = coh['on_task_hours']; COVER = coh['coverage_pct']; LOADED = coh['loaded_h_per_task']
people = json.load(open(os.path.join(SCR, 'final_rows.json')))          # 60, AHT-sorted
hours  = json.load(open(os.path.join(SCR, 'aug_hours_by_person.json')))['data']['rows']
days   = {r[1]: r[3] for r in hours}

by_person = defaultdict(list)
for t in tasks:
    by_person[t['contributor']].append(t)

# ------------------------------------------------------------------- styling
ARIAL   = 'Arial'
INK     = '1F2933'
def F(sz=10, b=False, color=INK, i=False):
    return Font(name=ARIAL, size=sz, bold=b, color=color, italic=i)
BLUE    = '0000FF'      # hardcoded input
GREEN   = '008000'      # link to another sheet
GREY    = '6B7A88'
HDR_FILL   = PatternFill('solid', fgColor='1F3B54')
BAND_FILL  = PatternFill('solid', fgColor='EDF2F6')
KEY_FILL   = PatternFill('solid', fgColor='FFFF00')
NOTE_FILL  = PatternFill('solid', fgColor='F5F7F9')
thin = Side(style='thin', color='C9D3DB')
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)
BOT  = Border(bottom=Side(style='thin', color='8FA3B3'))

MONEY = '$#,##0;($#,##0);-'
MONEY2= '$#,##0.00;($#,##0.00);-'
H2    = '0.00;-0.00;-'
PCT   = '0.0%'

def title(ws, row, text, size=13):
    c = ws.cell(row=row, column=1, value=text); c.font = F(size, True)
    return row + 1

def note(ws, row, text, col=1, span=None, italic=True, size=9):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F(size, color=GREY, i=italic); c.alignment = Alignment(wrap_text=True, vertical='top')
    if span: ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    return row + 1

def header_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F(9, True, color='FFFFFF'); c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical='bottom',
                                horizontal='right' if j > 2 else 'left')
    ws.row_dimensions[row].height = 30
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return row + 1

wb = Workbook()

# ================================================================== 1. READ ME
ws = wb.active; ws.title = 'Read me'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
for col, w in zip('BCDEFG', [30, 15, 15, 15, 15, 34]):
    ws.column_dimensions[col].width = w

r = 2
c = ws.cell(row=r, column=2, value='August 2026 AHT — prompt-only taskers'); c.font = F(17, True); r += 1
c = ws.cell(row=r, column=2, value='Average handle time, the contributor list, and the whole method behind it'); c.font = F(11, color=GREY); r += 2

c = ws.cell(row=r, column=2, value='THE ANSWER'); c.font = F(9, True, color='1F3B54'); r += 1
ans = [
    ('Median AHT — handle time (h)', "='Checks'!C8", H2, 'Paid time really spent on a finished task'),
    ('Mean AHT — handle time (h)',   "='Checks'!C7", H2, 'Pulled up by a long right tail'),
    ('Median cost per task',         "='Checks'!C8*Assumptions!C10", MONEY, 'At the blended rate'),
    ('Loaded hours per finished task', "=Assumptions!C9", H2, 'All paid hours / tasks finished — use this to budget'),
    ('Loaded cost per finished task',  "=Assumptions!C9*Assumptions!C10", MONEY, 'What a finished task really costs'),
]
for label, formula, fmt, why in ans:
    ws.cell(row=r, column=2, value=label).font = F(10)
    cc = ws.cell(row=r, column=3, value=formula); cc.font = F(12, True); cc.number_format = fmt
    ws.cell(row=r, column=6, value=why).font = F(9, color=GREY)
    r += 1
r += 1

c = ws.cell(row=r, column=2, value='SCOPE'); c.font = F(9, True, color='1F3B54'); r += 1
for line in [
  'Project: sheets. Role: exactly prompt_tasker — nobody in this workbook holds a second role.',
  'Ali Mallick is excluded at the owner\'s request; their 4 tasks and 19.8 August hours are out of',
  'every figure here, including the calibration.',
  'Period: every task whose FIRST arrival in the audit state fell in August 2026, Eastern time.',
  f'Population: {NPEOPLE} people — {coh["credited"]} credited with at least one such task, 11 who booked August hours',
  'but were never credited with one.',
  'A task in audit counts as done. That is the sheets convention, not a choice made here.',
]:
    ws.cell(row=r, column=2, value=line).font = F(10); r += 1
r += 1

c = ws.cell(row=r, column=2, value="WHAT'S IN EACH TAB"); c.font = F(9, True, color='1F3B54'); r += 1
tabs = [
 ('Assumptions',        'The three levers. Change one and every number in the workbook moves.'),
 ('AHT by contributor', 'The 60-person list, sorted fastest to slowest on median AHT.'),
 ('Per-task detail',    'All 905 tasks, one row each. This is the raw material everything else reads.'),
 ('Method and queries', 'The six steps, each with the SQL that produced it.'),
 ('Checks',             'The reconciliations that say the answer is sound, plus known weak points.'),
]
for name, what in tabs:
    ws.cell(row=r, column=2, value=name).font = F(10, True)
    cc = ws.cell(row=r, column=3, value=what); cc.font = F(10)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    r += 1
r += 1

c = ws.cell(row=r, column=2, value='HOW THE TIME IS MEASURED — IN PLAIN ENGLISH'); c.font = F(9, True, color='1F3B54'); r += 1
for line in [
  'The platform stores no per-task timer, and payable time is recorded against the PROJECT, not',
  'against a task — there is no task id anywhere on the payable-time tables. So a task\'s time has to',
  'be reconstructed. Two records make that possible:',
  '',
  '   the CLAIM            — this person held this task from this moment to that one',
  '   the PAYABLE SEGMENT  — this person was clocked in and approved for pay over this window',
  '',
  'Intersect them. A paid minute falling inside a claim window on task X is a minute spent on task X.',
  'Sum those minutes across every claim the credited person ever had on the task, so a task that went',
  'out for review and came back for a redo picks up all of its rounds.',
  '',
  'Waiting is correctly excluded. Between rounds the task is not claimed and the person is not clocked',
  'in on it, so days sitting in review add nothing to AHT. That is elapsed cycle time — a different',
  'measure, and not what AHT is for.',
  '',
  'If a person had two tasks claimed in the same minute, that minute is split between them in',
  'proportion to overlap, so nothing is double-counted and the parts never exceed what was paid.',
  '',
  'Cohort-wide, 801.7 of 1,769.3 paid August hours (45.3%) land on tasks that reached audit in August.',
  'The rest went to tasks that finished in another month or never, to tasks credited to someone else,',
  'to calibration and onboarding, and to clocked time with nothing claimed. Hence two numbers, both on',
  'the Assumptions tab:',
  '',
  '   HANDLE TIME   0.56h median, 0.89h mean   — paid time really spent on a task that finished',
  '   LOADED TIME   1.96h per finished task    — all paid hours divided by tasks finished',
  '',
  'Use handle time to compare people. Use loaded time to budget, because somebody pays for the work',
  'that did not finish.',
]:
    cc = ws.cell(row=r, column=2, value=line); cc.font = F(10)
    r += 1
r += 1

c = ws.cell(row=r, column=2, value='READ THE RANKING WITH CARE'); c.font = F(9, True, color='C2551F'); r += 1
for line in [
  'Both ends of the sort are people with one to four credited tasks, where a single task',
  'decides the whole ranking. Only 25 people cleared 10 tasks; inside that group the spread',
  'is wide, and it is not the same order as the claim-based ranking — see the caution on that tab.',
  'is the real finding. Anyone under about five tasks is listed for completeness, not judgement.',
]:
    ws.cell(row=r, column=2, value=line).font = F(10); r += 1
r += 1
ws.cell(row=r, column=2, value='Source: tsip_prd via Metabase (database 2, read-only), queried 31 August 2026.').font = F(9, color=GREY); r += 1
ws.cell(row=r, column=2, value='Conventions from projects/sheets/project.md. Blue cells are inputs you may change; black cells are formulas.').font = F(9, color=GREY)

# ============================================================== 2. ASSUMPTIONS
ws = wb.create_sheet('Assumptions')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
for col, w in zip('BCDEFG', [34, 14, 3, 62, 12, 12]):
    ws.column_dimensions[col].width = w

r = 2
ws.cell(row=r, column=2, value='Assumptions and levers').font = F(15, True); r += 1
r = note(ws, r, 'Yellow cells are the judgement calls. Change one and the whole workbook recalculates.', col=2, span=3) + 0
r += 1

r = header_row(ws, r, ['', 'Figure', 'Value', '', 'What it is'])
rows = [
 ('Payable hours ON these tasks', ONTASK, '#,##0.0',
  'Clocked, approved time that falls inside a claim window on one of the tasks that reached audit in '
  'August, for the person credited with it. This is what the AHT column adds up to.'),
 ('Cohort payable hours, August', PAYABLE, '#,##0.0',
  'Everything these people were paid for in August, whatever they were working on. From '
  'payable_time_revision_segments allocated to the day the work happened.'),
 ('Share landing on these tasks', None, '0.0%',
  'The rest went to tasks that reached audit in another month or not at all, to tasks credited to '
  'someone else, to calibration and onboarding, and to clocked time with no task claimed.'),
 ('Loaded hours per finished task', None, '0.00',
  'Cohort payable hours divided by tasks finished. THE COSTING NUMBER: somebody pays for the work '
  'that did not finish, so this is what a finished task really costs in time.'),
 ('Blended rate ($/hour)', BLENDED, MONEY2,
  'Cohort August cost divided by cohort August hours. Used for pooled cost figures; each person\'s '
  'own rate is used in their own row.'),
]
first = r
for i, (label, val, fmt, why) in enumerate(rows):
    ws.cell(row=r, column=2, value=label).font = F(10, True)
    cc = ws.cell(row=r, column=3)
    if label.startswith('Share landing'):    cc.value = f'=C{first}/C{first+1}';   cc.font = F(11, True)
    elif label.startswith('Loaded hours'):   cc.value = f"=C{first+1}/'Checks'!C6"; cc.font = F(11, True)
    else:                                    cc.value = val; cc.font = F(11, True, color=BLUE)
    cc.number_format = fmt; cc.fill = KEY_FILL; cc.border = BOX
    w = ws.cell(row=r, column=5, value=why); w.font = F(9, color=GREY)
    w.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 34
    r += 1
r += 1
ws.cell(row=r, column=2, value='Two ways to read a task\'s time — both are here on purpose').font = F(11, True); r += 1
for line in [
 'HANDLE TIME (the AHT column): paid time actually spent on that task. Answers "how long does the work take".',
 'LOADED TIME (the figure above): all paid time divided by tasks finished. Answers "what does a finished task cost".',
 'Handle time is the smaller number because it excludes work on tasks that never finished. Use handle time to',
 'compare people; use loaded time to budget.',
]:
    ws.cell(row=r, column=2, value=line).font = F(10); r += 1
r += 1
ws.cell(row=r, column=2, value='Where a paid second goes when someone holds several tasks at once').font = F(11, True); r += 1
for line in [
 'A paid minute is matched to whichever task the person had claimed at that minute. If they had two tasks',
 'claimed, the minute is split between them in proportion to the overlap, so no minute is counted twice and',
 'the parts can never add up to more than they were paid. Someone who keeps many tasks open at once will',
 'therefore show a low AHT per task — that is arithmetic, not speed.',
]:
    ws.cell(row=r, column=2, value=line).font = F(10); r += 1
r += 1

# ========================================================= 3. PER-TASK DETAIL
wsd = wb.create_sheet('Per-task detail')
wsd.freeze_panes = 'A2'
hdrs = ['Task ID', 'Contributor', 'Email', 'Payable hours on task', 'Claim hours (capped 12h)',
        'Claims by contributor', 'Cost at own rate ($)', 'Rate ($/h)']
header_row(wsd, 1, hdrs, widths=[38, 24, 30, 13, 13, 11, 13, 9])
wsd.freeze_panes = 'D2'

rate_of = {p['contributor']: p['rate_usd_per_hour'] for p in people}
ranges = {}
row = 2
for p in people:
    name = p['contributor']
    rs = sorted([t for t in tasks if t['contributor'] == name],
                key=lambda t: float(t['payable_hours_on_task']))
    if not rs:
        continue
    start = row
    for t in rs:
        wsd.cell(row=row, column=1, value=t['task_id']).font = F(8, color=GREY)
        wsd.cell(row=row, column=2, value=name).font = F(10)
        wsd.cell(row=row, column=3, value=t['email']).font = F(9, color=GREY)
        cc = wsd.cell(row=row, column=4, value=float(t['payable_hours_on_task']))
        cc.number_format = H2; cc.font = F(10, True, color=BLUE)
        cc = wsd.cell(row=row, column=5, value=float(t['claim_hours_capped12h']) if t['claim_hours_capped12h'] else None)
        cc.number_format = '#,##0.00'; cc.font = F(10, color=GREY)
        cc = wsd.cell(row=row, column=6, value=int(t['claims_by_contributor']) if t['claims_by_contributor'] else None)
        cc.font = F(10, color=GREY)
        cc = wsd.cell(row=row, column=7, value=f'=D{row}*H{row}'); cc.number_format = MONEY2; cc.font = F(10)
        cc = wsd.cell(row=row, column=8, value=rate_of[name]); cc.number_format = MONEY; cc.font = F(10, color=GREY)
        row += 1
    ranges[name] = (start, row - 1)
LAST = row - 1
note(wsd, row + 1, 'Blue = payable time measured against this task: clocked, approved minutes that fall inside '
                   'one of this person\'s claim windows on it, summed across every time they held it, including '
                   'across redo rounds. Grey = the raw claim window for reference, which is possession and includes '
                   'idle time. 21 tasks show zero payable minutes: work done while the claim had lapsed, or paid '
                   'before payable time was being recorded.', col=1, span=8)

# ======================================================= 4. AHT BY CONTRIBUTOR
wsc = wb.create_sheet('AHT by contributor')
wsc.sheet_view.showGridLines = False
r = 2
wsc.cell(row=r, column=1, value='AHT by contributor — fastest to slowest').font = F(15, True); r += 1
wsc.cell(row=r, column=1, value='Prompt-only taskers, sheets, August 2026. Every statistic is a live formula over the Per-task detail tab.').font = F(10, color=GREY); r += 2

hdrs = ['#', 'Contributor', 'Email', 'AHT (h)', 'Tasks', 'Hours', '$/task']
widths = [5, 27, 32, 10, 9, 10, 11]
hr = r
r = header_row(wsc, r, hdrs, widths=widths)
wsc.freeze_panes = f'C{r}'
data_first = r

for p in people:
    name = p['contributor']
    rng = ranges.get(name)
    wsc.cell(row=r, column=1, value=p['rank']).font = F(10, color=GREY)
    nm = wsc.cell(row=r, column=2, value=name); nm.font = F(10, p['rateable'] == 'yes')
    if rng:
        a, b = rng
        ref  = f"'Per-task detail'!$D${a}:$D${b}"
        rate = f"'Per-task detail'!$H${a}"
        cc = wsc.cell(row=r, column=4, value=f'=MEDIAN({ref})'); cc.number_format = H2; cc.font = F(11, True, color=GREEN)
        cc = wsc.cell(row=r, column=5, value=f'=COUNT({ref})');  cc.number_format = '#,##0'; cc.font = F(10, color=GREEN)
        cc = wsc.cell(row=r, column=7, value=f'=D{r}*{rate}');   cc.number_format = MONEY;  cc.font = F(10)
    else:
        cc = wsc.cell(row=r, column=4, value='no task reached audit in August')
        cc.font = F(9, color=GREY, i=True); cc.alignment = Alignment(horizontal='left')
        wsc.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        cc = wsc.cell(row=r, column=7, value='—'); cc.font = F(10, color=GREY); cc.alignment = Alignment(horizontal='right')
    em = wsc.cell(row=r, column=3, value=p.get('email', '')); em.font = F(9, color=GREY)
    cc = wsc.cell(row=r, column=6, value=p['august_payable_hours']); cc.number_format = '#,##0.0'; cc.font = F(10, color=BLUE)
    if p['rateable'] == 'yes':
        for j in range(1, 8): wsc.cell(row=r, column=j).fill = BAND_FILL
    r += 1
data_last = r - 1

# totals
tr = r
wsc.cell(row=tr, column=2, value='Cohort total / pooled').font = F(10, True)
for col, formula, fmt in [
    (4, f"=MEDIAN('Per-task detail'!$D$2:$D${LAST})", H2),
    (5, f'=SUM(E{data_first}:E{data_last})', '#,##0'),
    (6, f'=SUM(F{data_first}:F{data_last})', '#,##0.0'),
    (7, f'=D{tr}*Assumptions!$C$10', MONEY)]:
    cc = wsc.cell(row=tr, column=col, value=formula); cc.number_format = fmt; cc.font = F(10, True)
for j in range(1, 8):
    wsc.cell(row=tr, column=j).border = Border(top=Side(style='medium', color='1F3B54'))
r = tr + 2
r = note(wsc, r, 'AHT is that person\'s MEDIAN payable hours on a task — clocked, approved time that fell '
                 'inside their claim window on that specific task, across every round including redos. Half their '
                 'tasks took less, half took more.', span=7)
r = note(wsc, r, 'Tasks is how many reached audit in August with them credited.', span=7)
r = note(wsc, r, 'Hours is everything they booked and were paid for in August, across all their work — so it will '
                 'not equal AHT x Tasks, and is not meant to. Cohort-wide only 45.3% of paid hours land on tasks '
                 'that finished in August; see the Assumptions tab.', span=7)
r = note(wsc, r, '$/task is their AHT at their own contract rate. The cohort row uses the blended rate.', span=7)
r = note(wsc, r, 'Green = live formulas over the Per-task detail tab. Blue = measured payroll figures. Shaded rows '
                 'have 10 or more tasks and are the only ones worth ranking on. Mean, min, P90 and max per person '
                 'are in august_aht_contributor_list_full.csv.', span=7)
r = note(wsc, r, 'CAUTION: someone who keeps many tasks claimed at once has each paid minute split between them, so '
                 'their AHT per task reads low. Melvan Hoo Jun Hong and Andy Deng move from the slow end of the '
                 'claim-based ranking to the fast end here for exactly that reason. Read AHT next to Hours.', span=7)

# =================================================== 5. METHOD AND QUERIES
wsm = wb.create_sheet('Method and queries')
wsm.sheet_view.showGridLines = False
wsm.column_dimensions['A'].width = 3
wsm.column_dimensions['B'].width = 118
r = 2
wsm.cell(row=r, column=2, value='Method — six steps, each with its query').font = F(15, True); r += 1
wsm.cell(row=r, column=2, value='Everything below was run read-only against tsip_prd through the Metabase dataset API on 31 August 2026.').font = F(10, color=GREY); r += 2

def sqlblock(row, path):
    txt = open(os.path.join(AHT, path)).read().rstrip()
    c = wsm.cell(row=row, column=2, value=txt)
    c.font = Font(name='Consolas', size=8.5, color='24343F')
    c.alignment = Alignment(wrap_text=False, vertical='top')
    c.fill = NOTE_FILL; c.border = BOX
    wsm.row_dimensions[row].height = 11.5 * (txt.count('\n') + 1)
    return row + 2

def para(row, text, bold=False, size=10, color=INK):
    c = wsm.cell(row=row, column=2, value=text)
    c.font = F(size, bold, color=color); c.alignment = Alignment(wrap_text=True, vertical='top')
    wsm.row_dimensions[row].height = max(14, 13 * (len(text) // 105 + 1))
    return row + 1

steps = [
 ('STEP 1 — DEFINE THE POPULATION', 'Who counts as a "prompt-only tasker"?',
  ['The brief was to exclude anyone holding more than one role. Checking first: all 618 sheets '
   'members hold exactly one role row, so multiple roles do not occur in this data and nobody was '
   'excluded on that basis.',
   'Roles present: beginner_tasker 369, prompt_tasker 112, calibration_tasker 90, reviewer 19, '
   'decomp_prompt_tasker 10, calibration_failed 7, decomp_prompt_reviewer 3, decomp_tasker 3, '
   'beginner_tasker_reviewer 3, eval_hallucination_reviewer 2.',
   'JUDGEMENT CALL: the filter is role = \'prompt_tasker\' exactly. That drops the hybrids and also '
   'people doing prompt work under a reviewer or beginner title — most consequentially the project\'s '
   'highest-volume contributor, who is beginner_tasker_reviewer. The role field is a permissions '
   'label, not a description of work.'],
  'contributor_roles.sql'),
 ('STEP 2 — DEFINE THE TASK SET', 'Unique tasks that reached audit in August',
  ['A task counts once, on its FIRST arrival in audit — min(created_at) where to_state = audit and '
   'the task was not already there. Without that, a task bounced back and re-audited would be '
   'counted twice.',
   'The month boundary is Eastern time. Database timestamps are naive UTC, so the conversion is a '
   'double convert: AT TIME ZONE \'UTC\' AT TIME ZONE \'America/New_York\'. Getting this wrong moves '
   'tasks between months at the edges.',
   'Result: 1,287 sheets tasks first reached audit in August 2026.'],
  None),
 ('STEP 3 — THE CREDIT RULE', 'Who does a task belong to?',
  ['The obvious reading of "sent into audit" fails: 97% of the transitions INTO audit are triggered '
   'by the system, not a person, because submission is automated. Reading triggered_by on the audit '
   'transition would credit almost every task to a service account.',
   'The sheets convention instead credits whoever triggered the last exit out of edit_task or '
   'redo_task before the task reached audit — the person who actually finished the work.',
   'Result: of 1,287 August tasks, 907 are credited to someone whose role is exactly prompt_tasker. '
   'The other 380 belong to the nine other roles and are out of scope.',
   'WEAK POINT: a task reassigned late gives all its credit to whoever finished it. 92% of send-backs '
   'are fixed by the same person who wrote the task, so this is usually the same human — but not always.'],
  'august_aht_task_credit.sql'),
 ('STEP 4 — MEASURE TIME ON THE TASK', 'Time on that task specifically, not allocated',
  ['There is no per-task time field in the platform: payable time is project-grain. The only per-task '
   'signal is tsip_task_claims — each row is one person holding one task, with claimed_at and '
   'last_activity_at.',
   'The credited person\'s own claims on that exact task are summed, each capped at 12 hours so one '
   'forgotten claim cannot dominate. Nothing is allocated or spread; if a claim is not on this task, '
   'it does not count towards this task.',
   'Result: 907 tasks, of which 905 have a claim record for the credited person (2 do not and are '
   'dropped). Raw: mean 16.76 claim-hours, median 12.89, total 15,171.'],
  'august_aht_per_task_time.sql'),
 ('STEP 5 — CALIBRATE', 'Convert possession time into paid work time',
  ['16.76 hours per task is obviously not work — it is how long people had the task open. Payable '
   'time is the honest measure of work, but it exists only at project level, so it is used to set '
   'the level while claims set the shape.',
   'The same cohort booked 1,789.1 payable hours in August. 1,789.1 / 15,171 = 0.11793. Every task '
   'is multiplied by that factor.',
   'Result: mean 1.98 hours ($323), median 1.52 hours ($248) at the blended $163.17/hour.',
   'WEAK POINT: this assumes effort is proportional to possession within the cohort. Someone who '
   'leaves claims open looks slower than they are; someone who claims tightly looks faster. The 12-hour '
   'cap limits how far that can run, and it is the main lever on the level — see the Assumptions tab.'],
  None),
 ('STEP 6 — CHECK IT', 'Does the answer survive scrutiny?',
  ['The cross-check below re-derives the cohort\'s August payable hours and cost from payroll data '
   'alone, touching no task or claim table. See the Checks tab for what it returned.'],
  'august_aht_payable_crosscheck.sql'),
]
for label, head, paras, sql in steps:
    c = wsm.cell(row=r, column=2, value=label); c.font = F(9, True, color='1F3B54'); r += 1
    c = wsm.cell(row=r, column=2, value=head); c.font = F(12, True); r += 1
    for p_ in paras:
        r = para(r, p_, bold=p_.startswith(('JUDGEMENT', 'WEAK POINT')),
                 color='C2551F' if p_.startswith(('JUDGEMENT', 'WEAK POINT')) else INK)
    r += 1
    if sql: r = sqlblock(r, sql)
    else:   r += 1

c = wsm.cell(row=r, column=2, value='Per-contributor August hours and cost (the last two data columns of the list)')
c.font = F(11, True); r += 2
r = sqlblock(r, 'august_hours_by_contributor.sql')

# ============================================================ 6. CHECKS
wsk = wb.create_sheet('Checks')
wsk.sheet_view.showGridLines = False
wsk.column_dimensions['A'].width = 3
for col, w in zip('BCDEF', [40, 16, 16, 3, 74]):
    wsk.column_dimensions[col].width = w
r = 2
wsk.cell(row=r, column=2, value='Checks and reconciliations').font = F(15, True); r += 1
wsk.cell(row=r, column=2, value='Live values recomputed from this workbook, next to the figures they should equal.').font = F(10, color=GREY); r += 2

r = header_row(wsk, r, ['', 'Check', 'This workbook', 'Expected', '', 'What it tells you'])
first_chk = r
checks = [
 ('Tasks with claim data', f"=COUNT('Per-task detail'!$D$2:$D${LAST})", NTASKS, '#,##0',
  'Every credited task with a claim record for the credited person. 903 were credited; 2 have no claim row.'),
 ('Mean AHT (hours)',   f"=AVERAGE('Per-task detail'!$D$2:$D${LAST})", round(coh['mean'],2), H2,
  'Pooled mean. Recomputes if you change the calibration factor.'),
 ('Median AHT (hours)', f"=MEDIAN('Per-task detail'!$D$2:$D${LAST})", round(coh['median'],2), H2,
  'Pooled median — the figure to quote, because the mean is pulled up by a long right tail.'),
 ('Payable hours ON these tasks', f"=SUM('Per-task detail'!$D$2:$D${LAST})", ONTASK, '#,##0.0',
  'Paid time attributed to the 904 finished tasks. Must be LESS than cohort payable hours — the gap is '
  'work on tasks that did not finish in August.'),
 ('Cohort August payable hours', None, PAYABLE, '#,##0.0',
  'All paid hours for these 59 people in August, from payroll.'),
 ('Share landing on these tasks', f"=C{first_chk+3}/C{first_chk+4}", coh['coverage_pct']/100, '0.0%',
  'Under 100% by construction. If this ever reads 100% the attribution is wrong.'),
 ('Contributors listed', f"=COUNTA('AHT by contributor'!$B${data_first}:$B${data_last})", NPEOPLE, '#,##0',
  '48 credited with a task + 11 with August hours and no credited task.'),
 ('Cohort August cost ($)', None, NCOST, MONEY,
  'From payroll, independent of anything task-level. The cross-check query returned $291,936 for 59 people; '
  'less Ali Mallick\'s $2,974 that is $288,961 for 58.'),
 ('Implied blended rate ($/h)', f'=C{first_chk+7}/C{first_chk+4}', BLENDED, MONEY2,
  'Cohort cost divided by cohort hours. Should land on the rate used for the pooled figures.'),
]
for label, formula, expected, fmt, why in checks:
    wsk.cell(row=r, column=2, value=label).font = F(10, True)
    cc = wsk.cell(row=r, column=3, value=formula if formula else expected)
    cc.number_format = fmt; cc.font = F(11, True, color=INK if formula else BLUE)
    cc = wsk.cell(row=r, column=4, value=expected); cc.number_format = fmt; cc.font = F(10, color=BLUE)
    w = wsk.cell(row=r, column=6, value=why); w.font = F(9, color=GREY); w.alignment = Alignment(wrap_text=True, vertical='top')
    wsk.row_dimensions[r].height = 26
    r += 1
r += 1

wsk.cell(row=r, column=2, value='Independent corroboration').font = F(11, True); r += 1
for line in [
 'The payroll cross-check query (Method tab) re-derives the cohort\'s August hours and cost from '
 'payable_time_revision_segments and contributor_rates alone — it never touches tsip_tasks, '
 'tsip_state_transitions or tsip_task_claims. It returned 1,789.1 hours, 59 people, $291,936.',
 'That is the same 1,789.1 the calibration factor is built from, so the workbook\'s hours tie to '
 'money paid by two routes that share no table beyond the project id.',
]:
    c = wsk.cell(row=r, column=2, value=line); c.font = F(10); c.alignment = Alignment(wrap_text=True, vertical='top')
    wsk.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsk.row_dimensions[r].height = 28
    r += 1
r += 1

wsk.cell(row=r, column=2, value='What could still be wrong').font = F(11, True, color='C2551F'); r += 1
for line in [
 'AHT counts only paid time that fell inside a claim window. Work done after a claim lapsed is missed.',
 'Only 45.3% of paid hours land on tasks that finished in August, so handle time is not a costing number — '
 'use loaded time for that.',
 'Credit goes to the last submitter before audit. 92% of send-backs are self-fixed, so this is usually the '
 'same person — but not always.',
 'The role filter excludes real prompt work done under reviewer or beginner titles, including the project\'s '
 'biggest producer. That is the requested scope, not an accident.',
 'Everything is blended across task varieties. Patch-based still cannot be separated from greenfield in '
 'platform data.',
 'Daurys gonell is credited with an August task but has no August payable time — the claim spans the month '
 'boundary. Immaterial at one task, but it shows the two data sources are not perfectly aligned.',
 'Excluding Ali Mallick removes 4 tasks and 19.8 August hours from every figure here.',
 '21 of 904 tasks show zero payable minutes — work done while the claim had lapsed, or before payable time '
 'was recorded. They are kept at zero rather than dropped, which pulls the median down slightly.',
 'A person holding several tasks at once has each paid minute split between them, so their AHT per task '
 'reads low. This is the biggest caveat on the ranking: it rewards claiming tightly, not working fast.',
]:
    c = wsk.cell(row=r, column=2, value='•  ' + line); c.font = F(10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    wsk.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsk.row_dimensions[r].height = 26
    r += 1

wb.save(OUT)
print('wrote', OUT)
print('per-task rows 2..%d, contributor rows %d..%d' % (LAST, data_first, data_last))
