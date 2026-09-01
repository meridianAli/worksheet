"""Inject cached values next to every formula so the workbook renders in
previewers and Google Sheets, not only in Excel. Values are computed here from
the same source data the formulas reference; Excel still recalculates on open."""
import re, zipfile, shutil, json, csv, statistics as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

XL   = '/home/user/worksheet/aht/August-AHT-Analysis.xlsx'
SCR  = '/tmp/claude-0/-home-user-worksheet/968f285e-0587-5207-a956-c384e074891a/scratchpad/'
coh    = json.load(open(SCR + 'cohort.json'))
people = json.load(open(SCR + 'final_rows.json'))
tasks  = list(csv.DictReader(open('/home/user/worksheet/aht/august_aht_per_task.csv')))

wb = load_workbook(XL)
rate_of = {p['contributor']: p['rate_usd_per_hour'] for p in people}

vals = {}          # (sheetname, "D6") -> number
def put(sheet, col, row, v): vals[(sheet, f'{col}{row}')] = v

# --- Per-task detail: G = D * H
wsd = wb['Per-task detail']
per_row = {}
for r in range(2, wsd.max_row + 1):
    h = wsd.cell(row=r, column=4).value
    rate = wsd.cell(row=r, column=8).value
    if isinstance(h, (int, float)) and isinstance(rate, (int, float)):
        put('Per-task detail', 'G', r, h * rate)
        per_row[r] = h

# --- AHT by contributor
wsc = wb['AHT by contributor']
BLEND = round(coh['cost'] / coh['hours'], 2)
for r in range(1, wsc.max_row + 1):
    f = wsc.cell(row=r, column=4).value
    if not (isinstance(f, str) and f.startswith('=MEDIAN')):
        continue
    m = re.search(r"\$D\$(\d+):\$D\$(\d+)", f)
    a, b = int(m.group(1)), int(m.group(2))
    v = [per_row[i] for i in range(a, b + 1) if i in per_row]
    med = st.median(v)
    put('AHT by contributor', 'D', r, med)
    put('AHT by contributor', 'E', r, len(v))
    rate = wsd.cell(row=a, column=8).value
    put('AHT by contributor', 'G', r, med * rate)

allv = list(per_row.values())
# totals row: the one whose D holds the whole-range MEDIAN
for r in range(1, wsc.max_row + 1):
    f = wsc.cell(row=r, column=4).value
    if isinstance(f, str) and f.startswith('=MEDIAN') and ':$D$%d' % wsd.max_row in f:
        put('AHT by contributor', 'D', r, st.median(allv))
        put('AHT by contributor', 'E', r, len(allv))
        put('AHT by contributor', 'F', r, round(sum(p['august_payable_hours'] for p in people), 1))
        put('AHT by contributor', 'G', r, st.median(allv) * BLEND)

# --- Assumptions
put('Assumptions', 'C', 8, coh['on_task_hours'] / coh['payable'])
put('Assumptions', 'C', 9, coh['payable'] / len(allv))

# --- Checks
chk = {6: len(allv), 7: st.mean(allv), 8: st.median(allv),
       9: sum(allv), 11: coh['on_task_hours'] / coh['payable'],
       12: len(people), 14: coh['cost'] / coh['payable']}
for row, v in chk.items():
    put('Checks', 'C', row, v)

# --- Read me
rm = {7: st.median(allv), 8: st.mean(allv), 9: st.median(allv) * BLEND,
      10: coh['payable'] / len(allv), 11: coh['payable'] / len(allv) * BLEND}
wsr = wb['Read me']
for r in range(1, 30):
    f = wsr.cell(row=r, column=3).value
    if isinstance(f, str) and f.startswith('='):
        pass
sheet_order = wb.sheetnames

# map sheet name -> sheetN.xml
name_to_file = {}
for i, nm in enumerate(sheet_order, start=1):
    name_to_file[nm] = f'xl/worksheets/sheet{i}.xml'

tmp = XL + '.tmp'
zin = zipfile.ZipFile(XL)
zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
patched = 0
for item in zin.infolist():
    data = zin.read(item.filename)
    sheet = next((nm for nm, fn in name_to_file.items() if fn == item.filename), None)
    if sheet:
        xml = data.decode('utf-8')
        def repl(m):
            global patched
            ref = m.group(1)
            key = (sheet, ref)
            if key in vals and '<v>' not in m.group(0):
                v = vals[key]
                patched += 1
                return m.group(0)[:-len('</c>')] + f'<v>{v!r}</v></c>' if False else \
                       m.group(0).replace('</c>', f'<v>{v}</v></c>')
            return m.group(0)
        xml = xml.replace('<v />', '').replace('<v/>', '')
        xml = re.sub(r'<c r="([A-Z]+\d+)"[^>]*>(?:(?!</c>).)*?<f>(?:(?!</c>).)*?</c>', repl, xml)
        data = xml.encode('utf-8')
    zout.writestr(item, data)
zin.close(); zout.close()
shutil.move(tmp, XL)
print('cached values injected:', patched, 'of', len(vals), 'known')
