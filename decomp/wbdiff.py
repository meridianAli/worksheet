#!/usr/bin/env python3
"""Diff an input workbook against its output workbook, cell by cell.

Emits one record per change, classified so the coverage test can hold the
script to a different standard for each kind:

  hardcode  a literal value typed into a cell that was empty or different.
            A senior has to say these out loud - the analyst cannot derive them.
  formula   a formula added or changed. The script must make the analyst want
            the RESULT; it does not have to dictate the formula.
  label     a text/string cell - section headings, row captions, column heads.
  format    same value, different font/fill/number format/color.

Usage:
  python3 wbdiff.py --in <input.xlsx> --out <output.xlsx> --json changes.json
"""
import argparse, json, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

NUM = re.compile(r'^-?[\d,]*\.?\d+%?$')
# '=60000000' is a typed number wearing an equals sign, not a calculation.
LITERAL_FORMULA = re.compile(r'^=\s*\(?\s*(-?[\d,]*\.?\d+)\s*\)?\s*$')


def load(path):
    return (load_workbook(path, data_only=False, read_only=False),
            load_workbook(path, data_only=True, read_only=False))


def cellinfo(ws, wsv, coord):
    c = ws[coord]
    v = c.value
    cached = wsv[coord].value if wsv else None
    if v is None:
        return None
    if isinstance(v, ArrayFormula):
        # Compare array formulas by their text; the objects never compare equal.
        return {'kind': 'formula', 'formula': v.text, 'value': cached, 'array': True}
    if isinstance(v, str) and v.startswith('='):
        lit = LITERAL_FORMULA.match(v)
        if lit:
            return {'kind': 'hardcode', 'value': float(lit.group(1).replace(',', '')),
                    'as_formula': v}
        return {'kind': 'formula', 'formula': v, 'value': cached}
    if isinstance(v, str):
        return {'kind': 'label', 'value': v}
    return {'kind': 'hardcode', 'value': v}


def fmt(c):
    f = c.font
    return {
        'font': f.name, 'size': f.size, 'bold': f.bold,
        'color': (f.color.rgb if f.color and f.color.type == 'rgb' else None),
        'numfmt': c.number_format,
        'fill': (c.fill.fgColor.rgb if c.fill and c.fill.fgColor
                 and c.fill.fgColor.type == 'rgb' else None),
    }


def used(ws):
    return {c.coordinate for row in ws.iter_rows() for c in row
            if c.value is not None}


def diff(inp, outp):
    iw, ivw = load(inp)
    ow, ovw = load(outp)
    changes = []
    isheets, osheets = set(iw.sheetnames), set(ow.sheetnames)

    for name in ow.sheetnames:
        ows, ovs = ow[name], ovw[name]
        new_sheet = name not in isheets
        iws = ivs = None
        if not new_sheet:
            iws, ivs = iw[name], ivw[name]
        coords = used(ows) | (used(iws) if iws is not None else set())
        for coord in coords:
            after = cellinfo(ows, ovs, coord)
            before = None if new_sheet else cellinfo(iws, ivs, coord)
            if before == after:
                # same content - check formatting only where content exists
                if after is not None and not new_sheet:
                    fb, fa = fmt(iws[coord]), fmt(ows[coord])
                    if fb != fa:
                        changes.append({'sheet': name, 'cell': coord,
                                        'type': 'format', 'new_sheet': new_sheet,
                                        'before': fb, 'after': fa})
                continue
            if after is None:
                changes.append({'sheet': name, 'cell': coord, 'type': 'deleted',
                                'new_sheet': new_sheet, 'before': before, 'after': None})
                continue
            changes.append({'sheet': name, 'cell': coord, 'type': after['kind'],
                            'new_sheet': new_sheet, 'before': before, 'after': after,
                            'row': ows[coord].row, 'col': ows[coord].column,
                            'col_letter': get_column_letter(ows[coord].column)})
    for name in isheets - osheets:
        changes.append({'sheet': name, 'cell': None, 'type': 'sheet_removed'})
    return {
        'input': inp, 'output': outp,
        'sheets_added': sorted(osheets - isheets),
        'sheets_removed': sorted(isheets - osheets),
        'changes': changes,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--in', dest='inp', required=True)
    ap.add_argument('--out', dest='outp', required=True)
    ap.add_argument('--json', required=True)
    a = ap.parse_args()
    d = diff(a.inp, a.outp)
    op = gzip.open if a.json.endswith('.gz') else open
    with op(a.json, 'wt') as fh:
        json.dump(d, fh, indent=1, default=str)
    from collections import Counter
    c = Counter(x['type'] for x in d['changes'])
    print(f"{a.inp.split('/')[-1]} -> {a.outp.split('/')[-1]}")
    print(f"  sheets added: {d['sheets_added']}")
    print(f"  changes: {dict(c)}  total {len(d['changes'])}")


if __name__ == '__main__':
    main()
